# -*- coding: utf-8 -*-
"""生成《业务操作-资金流转勾稽矩阵.xlsx》（多 sheet，带样式）。
只读审计的交付物生成脚本，可复现。"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = r"E:\1target\SIEGPU\fundflow-audit"
os.makedirs(OUT_DIR, exist_ok=True)
OUT = os.path.join(OUT_DIR, "业务操作-资金流转勾稽矩阵.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="微软雅黑")
CELL_FONT = Font(size=10, name="微软雅黑")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SEV_FILL = {
    "高": PatternFill("solid", fgColor="FDEAEA"),
    "中": PatternFill("solid", fgColor="FFF4E0"),
    "低": PatternFill("solid", fgColor="EAF6EC"),
    "—": PatternFill("solid", fgColor="FFFFFF"),
}
SEV_FONT = {
    "高": Font(size=10, bold=True, color="C0392B", name="微软雅黑"),
    "中": Font(size=10, bold=True, color="B9770E", name="微软雅黑"),
    "低": Font(size=10, color="1E8449", name="微软雅黑"),
}


def sheet(wb, title, headers, rows, widths, sev_idx=None):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    for r in rows:
        ws.append(r)
    for ri in range(2, len(rows) + 2):
        for ci in range(1, len(headers) + 1):
            cell = ws.cell(row=ri, column=ci)
            cell.font = CELL_FONT
            cell.alignment = WRAP
            cell.border = BORDER
            if sev_idx is not None and ci == sev_idx + 1:
                key = str(rows[ri - 2][sev_idx]) if sev_idx < len(rows[ri - 2]) else "—"
                k = key[0] if key and key[0] in SEV_FILL else "—"
                cell.fill = SEV_FILL.get(k, SEV_FILL["—"])
                if k in SEV_FONT:
                    cell.font = SEV_FONT[k]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 26
    return ws


wb = Workbook()
wb.remove(wb.active)

# ---------- Sheet1 勾稽矩阵 ----------
sheet(wb, "勾稽矩阵(18动作)",
      ["#", "业务操作", "入口/位置", "资金流水", "池调整", "台账/计划同步", "科目映射", "勾稽结论", "严重度"],
      [
          [1, "付款（拆分）", "POST /payment-requests/{id}/disburse  payment_service.py:99-120", "✅ 每池一笔", "✅ 逐池校验余额", "核销行 Σ≤流水额", "✗", "✅ 勾稽", "—"],
          [2, "付款（非拆分）", "同上 :121-130", "✅ 单笔默认 OWN", "⚠️ 不校验余额", "同上", "✗", "⚠️ 部分", "中"],
          [3, "预付款落账", "POST /capital/prepayment  capital_service.py:336-371", "✅ 同事务双流水", "✅ 现金OUT+PREPAY IN", "✅ 台账同事务+共享幂等键", "✗", "✅ 勾稽", "—"],
          [4, "预付退回", "POST /capital/prepayment/refund :374-394", "✅ 双流水", "✅ PREPAY OUT+现金IN", "❌ 台账不更新", "✗", "🔴 断链", "高"],
          [5, "预付核销", "POST /capital/prepayment/offset :399-411", "✅ 单笔", "✅ PREPAY OUT", "❌ 台账不更新；前端不传发票", "✗", "🔴 断链", "高"],
          [6, "预付月结转", "settle_for_billing  prepayment_service.py:91-141", "✗ 不涉资金（设计）", "—（PREPAY 池不动）", "✅ 设备字段+台账双写", "✗", "⚠️ 口径矛盾", "中"],
          [7, "付款冲抵预付款", "_apply_prepayment_offset  payment_service.py:148-163", "✗ 现金流水已扣冲抵额", "实付=申请额−冲抵", "❌ 只写设备字段，不同步台账", "✗", "🔴 部分断链", "高"],
          [8, "金租放款（入池）", "disburse / add_disbursement  leasing_service.py:159-164,270-276", "✅ LEASING IN", "✅", "还款计划自动生成", "✗", "✅ 勾稽", "—"],
          [9, "金租放款（直付）", "add_disbursement :256-268", "✅ 负债IN+代付OUT", "✅ 净0", "计划照常", "❌ 科目常量未实现", "⚠️ 部分", "中"],
          [10, "置换归还（自动）", "execute_replacement  funding_service.py:13-89", "✅ 归还 IN（回填bank_id）", "✅ 净头寸剔除防双计", "replaced 增量防重", "✗", "✅ 勾稽", "—"],
          [11, "确认还款", "confirm_repayment  repayment_service.py:17-28", "❌ 不生成任何流水", "❌", "仅计划行实际值", "✗", "🔴 断链", "高"],
          [12, "金租利息", "—（全 services 确认）", "❌ 无任何资金动作", "❌", "仅计划/实际值字段", "✗", "🔴 断链", "高"],
          [13, "收款", "通用 POST /capital/transactions  capital.py:31-43", "✅ 默认 OWN IN", "⚠️ 跳过余额校验", "靠 reconcile_invoice 挂票", "✗", "⚠️ 部分", "中"],
          [14, "开票", "create_invoice  invoice_service.py:43-73", "✗ 纯票据台账（正确）", "—", "硬门4→自动出收入草稿", "收入侧有", "✅ 勾稽", "—"],
          [15, "收入确认审批", "on_approval_result  revenue_recognition_service.py:120-147", "✗ 不动资金池（正确）", "—", "voucher_json + 推 EBS", "✅ 唯一科目点", "✅ 勾稽", "—"],
          [16, "项目调配/归还", "allocate / return  capital_service.py:416-462,509-532", "✅ 2 笔 OUT/IN 净零", "⚠️ 净头寸校验非单池", "三表同事务", "✗", "⚠️ 部分", "中"],
          [17, "银行借款/还银行", "/bank-loan /repay-bank  capital_service.py:305-333", "✅ BANK IN/OUT", "✅ 授信+池余额双校验", "—", "✗", "✅ 勾稽", "—"],
          [18, "核销汇兑损益", "_book_fx_for_allocation / maybe_book_exchange_diff", "✅ 汇兑损益流水", "✅ 默认 OWN", "分摊到设备", "⚠️ 科目仅 note 文本", "⚠️ 专家漏列", "中"],
      ],
      [4, 20, 34, 18, 18, 26, 14, 12, 8], sev_idx=8)

# ---------- Sheet2 断链清单 ----------
sheet(wb, "断链清单(按严重度)",
      ["编号", "严重度", "问题", "位置", "后果", "修复建议"],
      [
          ["H1", "高", "确认还款不生成资金流水", "repayment_service.py:17-28", "还本付息现金流出不可见，池余额虚高，现金流测算失真", "confirm_repayment 同事务生成还本+付息流水，写 repayment.capital_transaction_id（列已预留）"],
          ["H2", "高", "金租利息无任何资金流水", "全 services 目录", "利息支出在资金池视角不存在", "补利息支出流水，或明确挂费用科目"],
          ["H3", "高", "预付退回/核销不更新 prepayments 台账行", "capital_service.py:374-411", "台账>池，对账块8必然标红；台账剩余虚高误导后续冲抵", "退回减 amount / 核销增 settled_amount，同步台账"],
          ["H4", "高", "预付核销不挂发票，字段可空", "CapitalView.vue:115、schemas/capital.py:116-117", "抵应付无票据勾稽；若带 invoice_id 又会混入 matched_amount 已付款口径", "前端加采购发票选择，后端改必填；区分已付 vs 预付抵付"],
          ["H5", "高", "付款冲抵只写设备字段不同步台账", "payment_service.py:148-163", "设备镜像与 prepayments 台账两套已结转口径分叉", "冲抵时同步台账 settled_amount"],
          ["H6", "高", "对账块8公式不区分结转/退回/核销", "reconciliation_service.py:264-312", "正常业务流下长期标红，告警失效", "按差异原因拆分三类展示"],
          ["M1", "中", "非拆分付款不校验池余额", "payment_service.py:121-130", "自有池可透支为负", "补 _assert_pool_sufficient"],
          ["M2", "中", "通用记一笔跳过余额校验", "capital_service.py:276-285", "可人为造成任意池透支", "OUT 方向加校验或二次确认"],
          ["M3", "中", "追加放款对同一验收无去重", "leasing_service.py:202-289", "双击→重复流水+重复还款计划", "同验收已放款金额≥验收金额时 409"],
          ["M4", "中", "销售验收放款血缘仅靠 project_id", "leasing_service.py:227-228", "错配项目无法拦截", "补订单/合同链式校验"],
          ["M5", "中", "项目调配按净头寸而非单池校验", "capital_service.py:434-437", "可能抽干单一池", "按池校验可调余额"],
          ["M6", "中", "直付负债科目常量未实现", "计划有、代码无", "直付负债入账无科目承接", "落地 LEASE_PAYABLE_ACCOUNT"],
          ["F1", "中", "核销汇兑损益流水被专家漏列", "payment_service.py:232-279、exchange_service.py:149-200", "会动 OWN 池+净头寸，未被纳入勾稽矩阵", "纳入矩阵，统一科目口径"],
          ["F2", "中", "repayments.capital_transaction_id 列空置", "models/repayment.py:24", "还款→流水关联设计好却未写，坐实 H1 为漏做", "confirm_repayment 回填该列"],
          ["F3", "中", "dim4 监管账户还款支出读永不生成的 source_type", "reconciliation_service.py:157", "留存余额=received−0，留存不足告警结构性失效", "改读还款流水真实来源"],
          ["F4", "中", "置换归还 IN 默认落 OWN 池的池间不对称", "funding_service.py:53-65", "OWN 池被抬高、BANK 池保留原 OUT，四池卡片失真", "明确置换归还落池规则"],
          ["F5", "中", "两条汇兑损益路径幂等键与科目行为不一致", "payment_service.py:246 vs exchange_service.py:167", "fx:{txn.id}:{inv.id} vs fx:{txn.id}；新版不查 ExchangeGainLossRule", "统一幂等键与科目落账"],
      ],
      [6, 6, 30, 34, 40, 40], sev_idx=1)

# ---------- Sheet3 资金池映射 ----------
sheet(wb, "资金池映射(动作→池→方向)",
      ["业务动作", "池", "方向", "source_type", "代码位置"],
      [
          ["银行借款", "BANK", "IN", "银行流贷", "capital_service.py:311-313"],
          ["还银行", "BANK", "OUT", "归还银行", "capital_service.py:327-329"],
          ["预付（现金出）", "from_pool(默认BANK)", "OUT", "预付", "capital_service.py:352-355"],
          ["预付（挂账）", "PREPAY", "IN", "预付", "capital_service.py:356-359"],
          ["预付退回（挂账出）", "PREPAY", "OUT", "预付", "capital_service.py:382-385"],
          ["预付退回（现金回）", "to_pool(默认BANK)", "IN", "预付", "capital_service.py:386-389"],
          ["预付核销", "PREPAY", "OUT", "预付", "capital_service.py:404-408"],
          ["付款/收款（拆分）", "LEASING/BANK/OWN", "OUT/IN", "金租融资/银行流贷/自有资金", "payment_service.py:77,113-120"],
          ["付款/收款（非拆分）", "OWN(默认)", "OUT/IN", "自有资金", "payment_service.py:122-130"],
          ["金租放款（入池）", "LEASING", "IN", "金租融资", "leasing_service.py:159-164,270-276"],
          ["金租直付（负债入账）", "LEASING", "IN", "金租融资", "leasing_service.py:256-261"],
          ["金租直付（代付货款）", "LEASING", "OUT", "金租融资", "leasing_service.py:263-268"],
          ["置换归还", "OWN(默认，净头寸剔除)", "IN", "归还流贷/归还自有", "funding_service.py:52-66"],
          ["项目调配", "OWN(默认)", "OUT/IN", "调配/调配归还", "capital_service.py:441-450,517-526"],
          ["核销汇兑损益", "OWN(默认)", "IN/OUT", "汇兑损益", "payment_service.py:258-266、exchange_service.py:180-192"],
          ["确认还款/利息", "—", "—", "不生成流水", "repayment_service.py:17-28"],
      ],
      [26, 22, 10, 24, 44])

# ---------- Sheet4 对账中心8块 ----------
sheet(wb, "对账中心8块",
      ["#", "勾稽块", "校验内容", "结论", "位置"],
      [
          [1, "销售全链路", "合同额→计费→开票→收款→收入，5 种差异 flag", "✅ 有断言", "reconciliation_service.py:34-68"],
          [2, "采购四单", "发票 vs 付款(matched_amount)、预付核销超额", "✅ 有断言", ":73-107"],
          [3, "资产交付", "采购≠入库、转固≠到货、点亮超转固", "✅ 有断言", ":115-139"],
          [4, "监管账户", "留存余额 < 下限", "⚠️ 还款支出读永不生成的 source_type=还款（F3）", ":144-168"],
          [5, "汇兑损益", "分摊不平 / 未分摊到设备", "✅ 有断言", ":173-192"],
          [6, "业财一致性", "SIEGPU vs EBS 四项对比", "⚠️ Mock 口径", ":197-230"],
          [7, "三流差异明细", "聚合块1+2 有 flag 的行", "⚠️ 仅展示无独立校验", ":235-260"],
          [8, "预付款勾稽", "PREPAY 池 ΣIN−ΣOUT vs 台账(amount−settled)", "⚠️ 有断言但公式缺陷（H6）", ":264-312"],
      ],
      [4, 14, 40, 32, 26])

# ---------- Sheet5 补充发现 ----------
sheet(wb, "补充发现(F1-F5)",
      ["编号", "发现", "证据", "影响", "建议"],
      [
          ["F1", "专家漏列『核销汇兑损益』这条真实资金流水", "payment_service._book_fx_for_allocation(:232-279) 与 exchange_service.maybe_book_exchange_diff(:149-200) 自动生成 source_type=汇兑损益 流水，默认 OWN 池、计入净头寸", "动 OWN 池+净头寸，直接影响对账块5与资金池余额，17 动作表未纳入", "纳入 18 动作矩阵，统一两条路径科目口径"],
          ["F2", "repayments.capital_transaction_id 列已预留却空置", "models/repayment.py:24 定义 FK，confirm_repayment 全文件无 CapitalTransaction 引用", "反证 H1 是『漏做』而非『设计豁免』", "confirm_repayment 回填该列"],
          ["F3", "dim4 监管账户还款支出读 source_type=还款", "reconciliation_service.py:157；全库仅 demo.py/tests 造该类型，生产永不生成", "留存余额=received−0，留存不足告警在还款侧结构性失效", "改读真实还款流水来源"],
          ["F4", "置换归还 IN 默认落 OWN 池", "funding_service.py:53-65 未设 pool（默认 OWN），原垫资付款在 BANK 池", "净头寸靠 _dir_sums 剔除防双计，但分池视角 OWN 被抬高、BANK 保留原 OUT", "明确置换归还落池规则"],
          ["F5", "两条汇兑损益路径幂等键/科目行为不一致", "payment_service.py:246(fx:{txn.id}:{inv.id}) vs exchange_service.py:167(fx:{txn.id})；新版不查 ExchangeGainLossRule", "幂等键格式不一致；科目码一条进 note、一条不进", "统一幂等键与科目落账"],
      ],
      [6, 34, 50, 34, 28])

wb.save(OUT)
print("SAVED:", OUT)
print("sheets:", wb.sheetnames)
