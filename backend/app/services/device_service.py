"""设备服务（一期 W1-2）：单台设备档案 CRUD + SN 生成 + 批次组合/移出 + Excel 批量导入 + 表外备查台账。

- SN 规则：GPU-{yyyymm}-{seq5}，缺省自动生成；显式 SN（导入场景）原样保留。
- status 物化列：W1-2 仅允许创建时写入，更新白名单不含 status（W3-4 起由设备状态机单点维护）。
- 批次纪律：同一台设备全局仅一条 active batch_devices 记录（service 校验 + 部分唯一索引兜底）；
  移出仅限上架前（订货/在途/到货/己方压测）；批次 flow_type 首次判定后固化，只升不降。
"""
from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.core.exceptions import BusinessError
from app.models.asset import Asset
from app.models.billing import Billing
from app.models.delivery import Order
from app.models.device import BatchDevice, Device, DeviceStage, OffBalanceRegister
from app.models.leasing import LeasingProcess
from app.models.master import EquipmentModel, Supplier
from app.models.project import Project
from app.utils.depreciation import depreciation_inputs
from app.utils.disbursement import disbursement_completion_pct, reached_threshold
from app.utils.ownership import derive_ownership
from app.utils.repayment_plan import add_months

# 批次移出允许的节点（seq<5 上架之前）
EARLY_STAGES = {"订货", "在途", "到货", "己方压测"}

# 设备 7 节点（一期 W3-4 设备粒度新路径）
DEVICE_STAGES = ["订货", "在途", "到货", "己方压测", "上架", "客户压测", "点亮验收"]
# 节点状态机：未开始→进行中；进行中→已完成/不合格；不合格→进行中（返工）；已完成→不合格（W5-6 返工）
DEVICE_STAGE_TRANSITIONS = {
    "未开始": {"进行中"},
    "进行中": {"已完成", "不合格"},
    "不合格": {"进行中"},
    "已完成": {"不合格"},
}
# 设备粒度 flow_type 集合（防双计闸用）——走 device_stages，禁走旧 6 节点
DEVICE_FLOW_TYPES = {"batch", "device", "transfer-resale"}

# Excel 导入支持的列
IMPORT_COLS = ["sn", "leasing_mode", "monthly_price", "purchase_value", "prepayment_amount", "ownership"]


def generate_sn(db: Session) -> str:
    """GPU-{yyyymm}-{seq5}：二期 W9-10 起委托 doc_number_service（规则表回迁，A8）。
    生成结果与一期硬编码完全一致：规则初始化从存量设备读当月最大 seq 接续（test_doc_number 锁死）。"""
    from app.services import doc_number_service
    return doc_number_service.generate_device_sn(db)


def create_device(db: Session, *, project_id, equipment_model_id, sn=None, order_id=None,
                  sales_contract_id=None, supplier_id=None, monthly_price=None, config=None,
                  leasing_mode=None, purchase_value=None, prepayment_amount=Decimal("0"),
                  ownership=None, operator_id=None) -> Device:
    proj = db.get(Project, project_id)
    if not proj or proj.deleted_at is not None:
        raise BusinessError("NOT_FOUND", "项目不存在", 404)
    if not db.get(EquipmentModel, equipment_model_id):
        raise BusinessError("NOT_FOUND", "设备型号不存在", 404)
    d = Device(
        sn=sn or generate_sn(db),
        project_id=project_id, equipment_model_id=equipment_model_id, order_id=order_id,
        sales_contract_id=sales_contract_id, supplier_id=supplier_id,
        monthly_price=monthly_price, config=config,
        leasing_mode=leasing_mode if leasing_mode is not None else proj.leasing_mode,  # 快照自项目
        purchase_value=purchase_value, prepayment_amount=prepayment_amount,
        status="订货",  # M-2：状态机唯一入口，创建恒为订货（schema 已不接受 status）
        ownership=ownership,
    )
    db.add(d)
    db.flush()
    from app.services import audit_service as _audit
    _audit.log(db, user_id=operator_id, action="CREATE", target_type="device",
               target_id=d.id, after_json={"sn": d.sn, "status": d.status})
    return d


def list_devices(db: Session, project_id=None, batch_id=None, status=None):
    stmt = select(Device).order_by(Device.created_at.desc())
    if project_id:
        stmt = stmt.where(Device.project_id == project_id)
    if batch_id:
        stmt = stmt.where(Device.batch_id == batch_id)
    if status:
        stmt = stmt.where(Device.status == status)
    return db.execute(stmt).scalars().all()


# v3.2 设备可租库存看板（F2）口径：仅「表内自有」设备参与自营出租。
# - 在租：已点亮验收 且 有未红冲按台计费（device 处于运营计费中）
# - 可租：已点亮验收 且 无未红冲计费（随时可下发租赁）
# - 待交付：尚未点亮（订货/在途/到货/己方压测/上架/客户压测）
# 金租/转售表外设备不参与自营出租，整体排除（与 _assert_light_rework_safe 的 Billing 存在性子查询同模式）。
PRE_LIT_STAGES = ["订货", "在途", "到货", "己方压测", "上架", "客户压测"]


def inventory_summary(db: Session) -> list[dict]:
    """按设备型号聚合表内自有设备的 可租/在租/待交付 数量。"""
    rented_ids = select(Billing.device_id).where(
        Billing.device_id.is_not(None),
        Billing.status != "已红冲",
        Billing.deleted_at.is_(None),
    ).distinct().subquery()
    is_rented = Device.id.in_(select(rented_ids.c.device_id))

    rows = db.execute(
        select(
            EquipmentModel.id, EquipmentModel.name, EquipmentModel.category,
            func.count(Device.id).label("total"),
            func.count(Device.id).filter(
                Device.status == "点亮验收", is_rented).label("rented"),
            func.count(Device.id).filter(
                Device.status == "点亮验收", ~is_rented).label("available"),
            func.count(Device.id).filter(
                Device.status.in_(PRE_LIT_STAGES)).label("pending"),
        )
        .join(Device, Device.equipment_model_id == EquipmentModel.id)
        .where(Device.deleted_at.is_(None), Device.ownership == "表内自有")
        .group_by(EquipmentModel.id, EquipmentModel.name, EquipmentModel.category)
        .order_by(EquipmentModel.name)
    ).all()

    return [{
        "model_id": str(mid), "model_name": name, "category": category,
        "total": total or 0, "rented": rented or 0,
        "available": available or 0, "pending": pending or 0,
    } for mid, name, category, total, rented, available, pending in rows]


def get_device_or_404(db: Session, device_id) -> Device:
    d = db.get(Device, device_id)
    if not d or d.deleted_at is not None:
        raise BusinessError("NOT_FOUND", "设备不存在", 404)
    return d


def update_device(db: Session, device_id, operator_id=None, **fields) -> Device:
    """白名单更新（schema DeviceUpdate 已约束字段集）；status/batch_id 不在其列。"""
    d = get_device_or_404(db, device_id)
    before = {k: str(getattr(d, k, None)) for k in fields}
    for k, v in fields.items():
        if v is not None and hasattr(d, k):
            setattr(d, k, v)
    db.flush()
    from app.services import audit_service as _audit
    _audit.log(db, user_id=operator_id, action="UPDATE", target_type="device",
               target_id=d.id, before_json=before,
               after_json={k: str(getattr(d, k, None)) for k in fields})
    return d


def delete_device(db: Session, device_id, operator_id=None) -> Device:
    d = get_device_or_404(db, device_id)
    d.deleted_at = datetime.now(timezone.utc)
    db.flush()
    from app.services import audit_service as _audit
    _audit.log(db, user_id=operator_id, action="DELETE", target_type="device",
               target_id=d.id, before_json={"sn": d.sn})
    return d


def _active_batch_row(db: Session, device_id) -> BatchDevice | None:
    return db.execute(
        select(BatchDevice).where(BatchDevice.device_id == device_id, BatchDevice.active.is_(True))
    ).scalars().first()


def add_to_batch(db: Session, *, device_id, batch_id, operator_id=None) -> BatchDevice:
    """设备挂入批次。守卫：批次订单存在；设备当前无 active 批次记录。"""
    d = get_device_or_404(db, device_id)
    batch = db.get(Order, batch_id)
    if not batch or batch.deleted_at is not None:
        raise BusinessError("NOT_FOUND", "批次订单不存在", 404)
    existing = _active_batch_row(db, device_id)
    if existing:
        raise BusinessError("DUPLICATE", "设备已在批次中，不能重复挂载", 409)
    if not batch.is_batch:
        batch.is_batch = True
    if batch.flow_type is None:  # 首次判定固化，只升不降
        batch.flow_type = "batch"
    bd = BatchDevice(batch_id=batch_id, device_id=device_id, action="加入", active=True,
                     operated_by=operator_id)
    db.add(bd)
    d.batch_id = batch_id
    db.flush()
    from app.services import audit_service as _audit
    _audit.log(db, user_id=operator_id, action="UPDATE", target_type="device",
               target_id=d.id, after_json={"batch_id": str(batch_id), "batch_action": "加入"})
    return bd


def remove_from_batch(db: Session, *, device_id, operator_id=None) -> BatchDevice:
    """设备移出批次。守卫：仅上架前节点允许（订货/在途/到货/己方压测）；flow_type 不回退。"""
    d = get_device_or_404(db, device_id)
    if d.status not in EARLY_STAGES:
        raise BusinessError("ILLEGAL_STATE", f"设备已进入「{d.status}」节点，不允许移出批次", 409)
    active = _active_batch_row(db, device_id)
    if not active:
        raise BusinessError("NOT_FOUND", "设备当前不在任何批次中", 404)
    active.active = False
    out = BatchDevice(batch_id=active.batch_id, device_id=device_id, action="移出", active=False,
                      operated_by=operator_id)
    db.add(out)
    d.batch_id = None
    db.flush()
    from app.services import audit_service as _audit
    _audit.log(db, user_id=operator_id, action="UPDATE", target_type="device",
               target_id=d.id, after_json={"batch_id": str(active.batch_id), "batch_action": "移出"})
    return out


def list_batch_devices(db: Session, batch_id=None, device_id=None):
    stmt = select(BatchDevice).order_by(BatchDevice.created_at.desc())
    if batch_id:
        stmt = stmt.where(BatchDevice.batch_id == batch_id)
    if device_id:
        stmt = stmt.where(BatchDevice.device_id == device_id)
    return db.execute(stmt).scalars().all()


def create_off_balance_register(db: Session, *, device_id, register_type, leasing_process_id=None,
                                start_date=None, end_date=None, note=None,
                                operator_id=None) -> OffBalanceRegister:
    get_device_or_404(db, device_id)
    r = OffBalanceRegister(
        device_id=device_id, register_type=register_type, leasing_process_id=leasing_process_id,
        start_date=start_date, end_date=end_date, note=note,
    )
    db.add(r)
    db.flush()
    from app.services import audit_service as _audit
    _audit.log(db, user_id=operator_id, action="CREATE", target_type="off_balance_register",
               target_id=r.id, after_json={"device_id": str(device_id), "register_type": register_type})
    return r


def list_off_balance_registers(db: Session, device_id=None):
    stmt = select(OffBalanceRegister).order_by(OffBalanceRegister.created_at.desc())
    if device_id:
        stmt = stmt.where(OffBalanceRegister.device_id == device_id)
    return db.execute(stmt).scalars().all()


# ============================ 一期 W5-6：一机一卡资产同步 ============================

# operation_status → 表外 register_type 映射（金租表外按 leasing_mode 细分、转售表外统一转售）
_OFF_BALANCE_REGISTER_TYPE = {
    "转售表外": "转售",
    "金租表外": "售后回租",  # 默认；leasing_mode='直租' 时下面覆盖为金租直租
}


def _ensure_off_balance_for_device(db: Session, device: Device, operator_id=None) -> OffBalanceRegister | None:
    """表外设备上架：写 off_balance_registers（不进 assets，避免污染折旧汇总）。幂等。"""
    existing = db.execute(select(OffBalanceRegister).where(
        OffBalanceRegister.device_id == device.id, OffBalanceRegister.deleted_at.is_(None)
    )).scalar_one_or_none()
    if existing is not None:
        return existing
    register_type = _OFF_BALANCE_REGISTER_TYPE.get(device.ownership or "", "金租直租")
    if device.ownership == "金租表外" and device.leasing_mode == "直租":
        register_type = "金租直租"
    return create_off_balance_register(db, device_id=device.id, register_type=register_type,
                                       operator_id=operator_id)


def _create_asset_card_for_device(db: Session, *, device: Device, operator_id=None) -> Asset:
    """上架→已转固未运营 资产卡（表内自有）。折旧字段暂 None，待点亮激活填。幂等。

    守卫：device.purchase_value 必须非空（单台原值是建卡 + 后续折旧的唯一输入）。
    """
    if device.purchase_value is None:
        raise BusinessError("BAD_REQUEST", f"设备 {device.sn} 缺 purchase_value，无法建资产卡", 400)
    existing = db.execute(select(Asset).where(
        Asset.device_id == device.id, Asset.deleted_at.is_(None)
    )).scalar_one_or_none()
    if existing is not None:
        return existing
    a = Asset(
        project_id=device.project_id, equipment_model_id=device.equipment_model_id,
        order_id=device.order_id, device_id=device.id, quantity=1,
        unit_original_value=device.purchase_value, total_original_value=device.purchase_value,
        residual_rate=Decimal("0.10"),  # 折旧字段全 None，待点亮填
        operation_status="已转固未运营", status="折旧中",
    )
    db.add(a)
    db.flush()
    from app.services import audit_service as _audit
    _audit.log(db, user_id=operator_id, action="CREATE", target_type="asset",
               target_id=a.id, after_json={"device_id": str(device.id), "operation_status": "已转固未运营"})
    return a


def _activate_asset_for_device(db: Session, *, device: Device, light_on_date: date,
                               operator_id=None) -> Asset:
    """点亮验收→运营中：填折旧字段 + start_date + operation_status='运营中'。幂等。

    若 device 跳过上架建卡直接进点亮（容错），现建一张并立即激活。
    """
    if device.purchase_value is None:
        raise BusinessError("BAD_REQUEST", f"设备 {device.sn} 缺 purchase_value，无法激活折旧", 400)
    a = db.execute(select(Asset).where(
        Asset.device_id == device.id, Asset.deleted_at.is_(None)
    )).scalar_one_or_none()
    if a is None:
        a = _create_asset_card_for_device(db, device=device, operator_id=operator_id)
    if a.operation_status == "运营中":
        return a  # 幂等：重复推进点亮不重算
    dep = depreciation_inputs(device.purchase_value)
    a.residual_value = dep["residual_value"]
    a.depreciable_value = dep["depreciable_value"]
    a.annual_depreciation = dep["annual_depreciation"]
    a.monthly_depreciation = dep["monthly_depreciation"]
    a.start_date = light_on_date
    a.end_date = add_months(light_on_date, dep["months"])
    a.operation_status = "运营中"
    db.flush()
    from app.services import audit_service as _audit
    _audit.log(db, user_id=operator_id, action="UPDATE", target_type="asset",
               target_id=a.id,
               after_json={"operation_status": "运营中", "start_date": str(light_on_date)})
    return a


def _sync_device_asset(db: Session, device: Device, stage: str, status: str,
                       actual_date: date | None, operator_id=None) -> None:
    """advance_device_stage 的资产同步派发（D1 两段式生命周期）。

    - 上架→已完成：表内自有建资产卡（已转固未运营）；表外走 off_balance_registers。
    - 点亮验收→已完成：表内自有点亮激活（起折旧）；表外无动作（不计提折旧）。
    其余节点不动资产。ownership 未设置时不建卡（数据质量由导入/校验把关，非本函数职责）。
    """
    if stage == "上架" and status == "已完成":
        # W7-8 D1 settle_ownership：ownership 为 None 时由 leasing_mode 派生（仅填 None，
        # 显式入参永远优先 → 49+9 现有设备测试零回归）。售后回租 → 表内自有（先转固，非表外）。
        if device.ownership is None:
            device.ownership = derive_ownership(device.leasing_mode)
            if device.ownership is not None:
                db.flush()
        if device.ownership == "表内自有":
            _create_asset_card_for_device(db, device=device, operator_id=operator_id)
        elif device.ownership in ("金租表外", "转售表外"):
            _ensure_off_balance_for_device(db, device, operator_id=operator_id)
    elif stage == "点亮验收" and status == "已完成" and actual_date is not None:
        if device.ownership == "表内自有":
            _activate_asset_for_device(db, device=device, light_on_date=actual_date,
                                       operator_id=operator_id)


# ============================ 一期 W3-4：设备状态机 ============================

def _derive_device_status(stages: list[DeviceStage]) -> str:
    """纯函数：由节点行派生 device.status 物化列。

    stages 须按 seq 升序。返回首个未「已完成」的 stage 名；全部已完成→点亮验收；无行→订货。
    「不合格」算未完成，device.status 停在该节点名（细节读行）。
    """
    if not stages:
        return "订货"
    for s in stages:
        if s.status != "已完成":
            return s.stage
    return "点亮验收"


def list_device_stages(db: Session, device_id) -> list[DeviceStage]:
    return db.execute(
        select(DeviceStage)
        .where(DeviceStage.device_id == device_id, DeviceStage.deleted_at.is_(None))
        .order_by(DeviceStage.seq)
    ).scalars().all()


def _ensure_device_stages(db: Session, device_id) -> list[DeviceStage]:
    """懒初始化：缺则建 7 行未开始；已存在直接返回（按 seq 升序）。幂等。"""
    existing = list_device_stages(db, device_id)
    if existing:
        return existing
    rows = [DeviceStage(device_id=device_id, stage=st, seq=i, status="未开始")
            for i, st in enumerate(DEVICE_STAGES, 1)]
    db.add_all(rows)
    db.flush()
    return rows


def init_device_stages(db: Session, device_id, operator_id=None) -> list[DeviceStage]:
    """显式初始化设备 7 节点（UI/批量可调；未进交付的设备无节点行）。"""
    get_device_or_404(db, device_id)
    return _ensure_device_stages(db, device_id)


def _assert_light_rework_safe(db: Session, device_id) -> None:
    """D5：点亮验收 已完成→不合格 返工前置守门。

    已有运营中资产或按台计费 → 抛 STATE_ERROR（财务副作用不可静默回退，须先红冲计费 + 处置资产）。
    表外设备（金租/转售）点亮验收不建资产 → 可自由返工；表内自有点亮即建卡 → 被拦。
    """
    has_asset = db.execute(
        select(Asset.id).where(
            Asset.device_id == device_id, Asset.operation_status == "运营中",
            Asset.deleted_at.is_(None),
        ).limit(1)
    ).scalar_one_or_none() is not None
    has_billing = db.execute(
        select(Billing.id).where(
            Billing.device_id == device_id, Billing.deleted_at.is_(None),
        ).limit(1)
    ).scalar_one_or_none() is not None
    if has_asset or has_billing:
        raise BusinessError(
            "STATE_ERROR",
            "该设备已点亮验收并建卡/计费，返工请先红冲按台计费与处置资产",
            409,
        )


def advance_device_stage(db: Session, *, device_id, stage, status,
                         actual_date=None, attachment_path=None, notes=None,
                         operator_id=None):
    """推进设备单节点：校验状态机→更新行→重算 device.status 物化列→同步批次聚合状态。

    返回 (device, stage_row)。点亮验收→已完成 在 W3-4 仅置完成（按台资产/计费在 W5-6）。
    """
    d = get_device_or_404(db, device_id)
    if stage not in DEVICE_STAGES:
        raise BusinessError("BAD_REQUEST", f"未知设备节点：{stage}", 400)
    stages = _ensure_device_stages(db, d.id)
    row = next((s for s in stages if s.stage == stage), None)
    if row is None:
        raise BusinessError("NOT_FOUND", f"设备 {d.sn} 无节点 {stage}", 404)
    allowed = DEVICE_STAGE_TRANSITIONS.get(row.status, set())
    if status not in allowed:
        raise BusinessError("ILLEGAL_TRANSITION", f"节点 {stage} 不允许 {row.status} → {status}", 409)
    # D5：点亮验收 已完成→不合格 返工守门（有运营中资产或按台计费 → 必须先红冲/处置）
    if stage == "点亮验收" and row.status == "已完成" and status == "不合格":
        _assert_light_rework_safe(db, d.id)
    before = row.status
    row.status = status
    if actual_date is not None:
        row.actual_date = actual_date
    if attachment_path is not None:
        row.attachment_path = attachment_path
    if notes is not None:
        row.notes = notes
    d.status = _derive_device_status(stages)  # stages 已就地变更，直接派生
    db.flush()
    # W5-6：上架建卡 / 点亮激活 的资产同步（D1 两段式生命周期）
    _sync_device_asset(db, d, stage, status, actual_date, operator_id=operator_id)
    # 二期 W7-8：自动投保 hook（advisory——无配置/异常只记日志，绝不阻塞设备推进）
    try:
        from app.services import insurance_service as _ins
        if stage == "在途" and status in ("进行中", "已完成") and d.batch_id:
            _ins.maybe_auto_transport_policy(db, batch_id=d.batch_id, operator_id=operator_id)
        elif stage == "点亮验收" and status == "已完成":
            _ins.maybe_auto_property_policy(db, device=d, operator_id=operator_id)
    except Exception:  # noqa: BLE001
        import logging
        logging.getLogger(__name__).warning("insurance auto-policy hook failed: device=%s stage=%s", d.id, stage)
    if d.batch_id:
        _sync_batch_status(db, d.batch_id)
        # W7-8 D3：点亮验收完成 → 批次放款达成率达阈值 → 自动建金租融资申请（幂等哨兵防二建）
        if stage == "点亮验收" and status == "已完成":
            _maybe_trigger_disbursement_todo(db, d.batch_id, operator_id)
    from app.services import audit_service as _audit
    _audit.log(db, user_id=operator_id, action="UPDATE", target_type="device",
               target_id=d.id,
               before_json={"stage": stage, "status": before},
               after_json={"stage": stage, "status": status, "device_status": d.status})
    from app.services import workflow_service as _wf
    _wf.after_action(db, d.project_id)
    return d, row


def advance_batch_stages(db: Session, *, batch_id, stage, status,
                         actual_date=None, attachment_path=None, notes=None,
                         operator_id=None) -> dict:
    """批量推进：批内所有 active 设备推进同一节点。返回 {ok, fail}（失败=状态机拒绝等）。"""
    dev_ids = [bd.device_id for bd in db.execute(
        select(BatchDevice).where(
            BatchDevice.batch_id == batch_id,
            BatchDevice.active.is_(True),
            BatchDevice.deleted_at.is_(None),
        )
    ).scalars().all()]
    ok = fail = 0
    for did in dev_ids:
        try:
            # W5-6 HIGH 修复：每台用 SAVEPOINT 隔离。_sync_device_asset 可能在 row.status 已 flush
            # 之后抛错（如表内自有设备缺 purchase_value），不隔离会让失败台的“已完成”节点随端点
            # commit 落库却无资产卡（“完成无卡”悬挂态，无法重推）。begin_nested 失败回滚到 savepoint
            # （只回滚这台的 flush，含 audit/after_action），成功则释放——批内其余设备不受影响。
            with db.begin_nested():
                advance_device_stage(db, device_id=did, stage=stage, status=status,
                                     actual_date=actual_date, attachment_path=attachment_path,
                                     notes=notes, operator_id=operator_id)
            ok += 1
        except BusinessError:
            fail += 1
    return {"ok": ok, "fail": fail}


def _aggregate_batch_status(db: Session, batch_id) -> str | None:
    """批内设备状态聚合：全点亮验收→已点亮；否则取瓶颈（进度最靠前设备所在节点名）。"""
    devs = list_devices(db, batch_id=batch_id)
    if not devs:
        return None
    if all(d.status == "点亮验收" for d in devs):
        return "已点亮"
    seqs = [DEVICE_STAGES.index(d.status) for d in devs if d.status in DEVICE_STAGES]
    return DEVICE_STAGES[min(seqs)] if seqs else "订货"


def _sync_batch_status(db: Session, batch_id):
    """把聚合状态写入 orders.batch_status 独立字段（不复用 orders.status，审计 A3）。"""
    batch = db.get(Order, batch_id)
    if batch is not None:
        batch.batch_status = _aggregate_batch_status(db, batch_id)
        db.flush()


# ============================ 一期 W7-8：放款条件联动 ============================

def _batch_light_completion(db: Session, batch_id) -> tuple[int, int]:
    """批次点亮达成率（lit, total）。total=批内未删设备数；lit=其中点亮验收已完成的。

    与 _aggregate_batch_status（字符串聚合）正交：纯数值派生，不存储（D3）。
    返工（已完成→不合格）会让该设备点亮行离开 已完成 → lit 自然下降。
    """
    dev_ids = [r[0] for r in db.execute(
        select(Device.id).where(
            Device.batch_id == batch_id, Device.deleted_at.is_(None)
        )
    ).all()]
    total = len(dev_ids)
    if total == 0:
        return 0, 0
    lit = db.execute(
        select(func.count()).select_from(DeviceStage).where(
            DeviceStage.device_id.in_(dev_ids),
            DeviceStage.stage == "点亮验收",
            DeviceStage.status == "已完成",
            DeviceStage.deleted_at.is_(None),
        )
    ).scalar_one()
    return lit, total


def _resolve_project_leasing_supplier(db: Session, project_id):
    """解析放款金租机构：①项目最近 leasing_process.supplier_id；②否则 is_leasing_org 首条；③都无→None。"""
    recent = db.execute(
        select(LeasingProcess.supplier_id).where(
            LeasingProcess.project_id == project_id, LeasingProcess.deleted_at.is_(None)
        ).order_by(LeasingProcess.created_at.desc()).limit(1)
    ).first()
    if recent is not None:
        return recent[0]
    s = db.execute(
        select(Supplier.id).where(
            Supplier.is_leasing_org.is_(True), Supplier.deleted_at.is_(None)
        ).order_by(Supplier.created_at.asc()).limit(1)
    ).first()
    return s[0] if s is not None else None


def _ensure_disbursement_leasing_process(db: Session, batch: Order, operator_id=None):
    """达阈值自动建金租融资申请（D4 零改复用 leasing_service.create_process）。

    - 金租机构缺失 → 跳过 + 记审计（靠向导 get_my_tasks 待办暴露，操作员手补 supplier）。
    - total_amount = Σ 批内 devices.purchase_value；financing_type 按 proj.leasing_mode 派生。
    返回 LeasingProcess 或 None（跳过）。
    """
    from app.services import audit_service as _audit
    from app.services import leasing_service as _lsvc

    supplier_id = _resolve_project_leasing_supplier(db, batch.project_id)
    if supplier_id is None:  # M4：无金租机构静默跳过（不抛错、不阻塞点亮完成）
        _audit.log(db, user_id=operator_id, action="UPDATE", target_type="order",
                   target_id=batch.id,
                   after_json={"disbursement_skip_reason": "no_funding_supplier"})
        return None
    total_amount = db.execute(
        select(func.coalesce(func.sum(Device.purchase_value), 0)).where(
            Device.batch_id == batch.id, Device.deleted_at.is_(None)
        )
    ).scalar_one() or Decimal("0")
    proj = db.get(Project, batch.project_id)
    leasing_mode = proj.leasing_mode if proj is not None else None
    financing_type = "金租直租" if leasing_mode == "直租" else "金租回租"
    proc = _lsvc.create_process(db, project_id=batch.project_id, supplier_id=supplier_id,
                                total_amount=total_amount, leasing_mode=leasing_mode,
                                financing_type=financing_type)
    return proc


def _maybe_trigger_disbursement_todo(db: Session, batch_id, operator_id=None):
    """点亮完成钩子：达阈值且未建过 → 自动建 leasing_process 并写幂等哨兵。

    哨兵 = orders.disbursement_todo_process_id（首次达阈值写一次）。
    返工后 pct 下降但哨兵不回退（leasing_process 是已落单业务单据，走审批驳回不机械删——不对称已记录）。
    """
    batch = db.get(Order, batch_id)
    if batch is None or batch.deleted_at is not None:
        return
    if batch.disbursement_todo_process_id is not None:  # 幂等：已建过直接返回
        return
    lit, total = _batch_light_completion(db, batch_id)
    if not reached_threshold(lit, total, batch.disbursement_threshold_pct):
        return
    proc = _ensure_disbursement_leasing_process(db, batch, operator_id=operator_id)
    if proc is not None:
        batch.disbursement_todo_process_id = proc.id
        db.flush()


def resolve_flow_type(db: Session, order: Order) -> str | None:
    """判定订单交付路径。只升不降：flow_type 已固化直接返回；否则以 devices.batch_id 关联为准。

    返回 None=旧 6 节点路径；返回 batch/device/transfer-resale=设备粒度路径。
    """
    if order.flow_type is not None:
        return order.flow_type
    linked = db.execute(
        select(Device.id).where(Device.batch_id == order.id, Device.deleted_at.is_(None)).limit(1)
    ).scalar_one_or_none()
    if linked is not None:
        order.flow_type = "batch"  # 设备挂入即视为批次载体，固化
        db.flush()
        return order.flow_type
    # M-1：单台订单经 order_id 直连设备（非批次挂载）→ device 路径
    attached = db.execute(
        select(Device.id).where(Device.order_id == order.id, Device.deleted_at.is_(None)).limit(1)
    ).scalar_one_or_none()
    if attached is not None:
        order.flow_type = "device"
        db.flush()
        return order.flow_type
    return None


def assert_legacy_path(db: Session, order: Order) -> None:
    """防双计闸（discipline ②）：设备粒度订单禁止走旧 6 节点入口（advance_stage/light_on/generate_billing）。

    is_batch 是硬信号，独立于设备是否已挂（防 batch 订单在设备挂入前调旧入口致 500）。
    """
    if order.is_batch or resolve_flow_type(db, order) in DEVICE_FLOW_TYPES:
        raise BusinessError(
            "FLOW_TYPE_DEVICE",
            "设备粒度订单请走设备路径（device_stages），不支持旧 6 节点操作",
            409,
        )


def import_devices(db: Session, *, project_id, equipment_model_id, filebytes: bytes,
                   operator_id=None) -> int:
    """Excel 批量导入设备（复用 excel_service 模式）：每行建一台设备，缺省自动 SN。

    支持列：sn / leasing_mode / monthly_price / purchase_value / prepayment_amount / ownership
    """
    proj = db.get(Project, project_id)
    if not proj or proj.deleted_at is not None:
        raise BusinessError("NOT_FOUND", "项目不存在", 404)
    if not db.get(EquipmentModel, equipment_model_id):
        raise BusinessError("NOT_FOUND", "设备型号不存在", 404)
    wb = load_workbook(BytesIO(filebytes), read_only=True)
    ws = wb.active
    headers = [str(c.value).strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(v is None for v in row):
            continue
        data = {}
        for i, h in enumerate(headers):
            if h in IMPORT_COLS and i < len(row) and row[i] not in (None, ""):
                data[h] = row[i]
        for k in ("monthly_price", "purchase_value", "prepayment_amount"):
            if k in data:
                data[k] = Decimal(str(data[k]))
        create_device(db, project_id=project_id, equipment_model_id=equipment_model_id,
                      operator_id=operator_id, **data)
        count += 1
    return count
