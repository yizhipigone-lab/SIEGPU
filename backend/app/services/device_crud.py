"""设备档案 CRUD + SN 生成 + 库存看板 + Excel 批量导入（自 device_service 拆出）。

- SN 规则：GPU-{yyyymm}-{seq5}，缺省自动生成；显式 SN（导入场景）原样保留。
- status 物化列：仅允许创建时写入，更新白名单不含 status（W3-4 起由设备状态机单点维护）。
"""
from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.core.exceptions import BusinessError
from app.models.billing import Billing
from app.models.device import Device, DeviceStage
from app.models.master import EquipmentModel
from app.models.project import Project

# Excel 导入支持的列（缺陷#11：中文表头优先，英文旧表头兼容）
IMPORT_COLS = ["sn", "leasing_mode", "monthly_price", "purchase_value", "prepayment_amount", "ownership"]
# 中文表头 → 内部字段名（导入模板已改中文，旧英文模板继续可用）
_HEADER_MAP = {
    "SN 序列号": "sn", "SN": "sn", "sn": "sn",
    "金租模式": "leasing_mode", "leasing_mode": "leasing_mode",
    "单台月计费额(元)": "monthly_price", "monthly_price": "monthly_price",
    "单台采购原值(元)": "purchase_value", "purchase_value": "purchase_value",
    "预付款分摊(元)": "prepayment_amount", "prepayment_amount": "prepayment_amount",
    "权属": "ownership", "ownership": "ownership",
}


def generate_sn(db: Session) -> str:
    """GPU-{yyyymm}-{seq5}：二期 W9-10 起委托 doc_number_service（规则表回迁，A8）。
    生成结果与一期硬编码完全一致：规则初始化从存量设备读当月最大 seq 接续（test_doc_number 锁死）。"""
    from app.services import doc_number_service
    return doc_number_service.generate_device_sn(db)


def create_device(db: Session, *, project_id, equipment_model_id, sn=None, order_id=None,
                  sales_contract_id=None, supplier_id=None, monthly_price=None, config=None,
                  leasing_mode=None, purchase_value=None, prepayment_amount=Decimal("0"),
                  prepayment_date=None, ownership=None, operator_id=None) -> Device:
    proj = db.get(Project, project_id)
    if not proj or proj.deleted_at is not None:
        raise BusinessError("NOT_FOUND", "项目不存在", 404)
    if not db.get(EquipmentModel, equipment_model_id):
        raise BusinessError("NOT_FOUND", "设备型号不存在", 404)
    d = Device(
        sn=sn or generate_sn(db),
        project_id=project_id, equipment_model_id=equipment_model_id, order_id=order_id,
        sales_contract_id=sales_contract_id, supplier_id=supplier_id,
        monthly_price=monthly_price, config=config,
        leasing_mode=leasing_mode if leasing_mode is not None else proj.leasing_mode,  # 快照自项目
        purchase_value=purchase_value, prepayment_amount=prepayment_amount,
        prepayment_date=prepayment_date,
        status="订货",  # M-2：状态机唯一入口，创建恒为订货（schema 已不接受 status）
        ownership=ownership,
    )
    db.add(d)
    db.flush()
    # 缺陷#16：建档即视为「已订货」——自动建 7 节点行并置订货已完成，
    # 避免用户跳节点推进（在途/到货）后状态卡「订货」。
    from app.services.device_stage_machine import DEVICE_STAGES
    stage_rows = [DeviceStage(device_id=d.id, stage=st, seq=i, status="已完成" if st == "订货" else "未开始")
                  for i, st in enumerate(DEVICE_STAGES, 1)]
    db.add_all(stage_rows)
    db.flush()
    from app.services import audit_service as _audit
    _audit.log(db, user_id=operator_id, action="CREATE", target_type="device",
               target_id=d.id, after_json={"sn": d.sn, "status": d.status})
    # S3（缺陷#6）：设备登记预付款 → 自动落台账行（供应商/日期可空待补）
    from app.services import prepayment_service as _pp
    _pp.ensure_device_ledger(db, d, actor_id=operator_id)
    return d


def list_devices(db: Session, project_id=None, batch_id=None, status=None):
    stmt = select(Device).order_by(Device.created_at.desc())
    if project_id:
        stmt = stmt.where(Device.project_id == project_id)
    if batch_id:
        stmt = stmt.where(Device.batch_id == batch_id)
    if status:
        stmt = stmt.where(Device.status == status)
    return db.execute(stmt).scalars().all()


# v3.2 设备可租库存看板（F2）口径：仅「表内自有」设备参与自营出租。
# - 在租：已点亮验收 且 有未红冲按台计费（device 处于运营计费中）
# - 可租：已点亮验收 且 无未红冲计费（随时可下发租赁）
# - 待交付：尚未点亮（订货/在途/到货/己方压测/上架/客户压测）
# 金租/转售表外设备不参与自营出租，整体排除（与 _assert_light_rework_safe 的 Billing 存在性子查询同模式）。
PRE_LIT_STAGES = ["订货", "在途", "到货", "己方压测", "上架", "客户压测"]


def inventory_summary(db: Session) -> list[dict]:
    """按设备型号聚合设备的 可租/在租/待交付 数量（缺陷#13：分两个口径）。

    - 自营（表内自有）：可对外出租的自有设备（原口径不变）
    - 金租租入（金租表外）：金租直租/售后回租租入、同样可对客户出租的设备，
      单列展示供参考，不与自营混算（权属与折旧口径不同）。
    """
    rented_ids = select(Billing.device_id).where(
        Billing.device_id.is_not(None),
        Billing.status != "已红冲",
        Billing.deleted_at.is_(None),
    ).distinct().subquery()
    is_rented = Device.id.in_(select(rented_ids.c.device_id))

    rows = db.execute(
        select(
            EquipmentModel.id, EquipmentModel.name, EquipmentModel.category,
            Device.ownership.label("ownership"),
            func.count(Device.id).label("total"),
            func.count(Device.id).filter(
                Device.status == "点亮验收", is_rented).label("rented"),
            func.count(Device.id).filter(
                Device.status == "点亮验收", ~is_rented).label("available"),
            func.count(Device.id).filter(
                Device.status.in_(PRE_LIT_STAGES)).label("pending"),
        )
        .join(Device, Device.equipment_model_id == EquipmentModel.id)
        .where(Device.deleted_at.is_(None),
               Device.ownership.in_(["表内自有", "金租表外"]))  # 缺陷#13：金租租入纳入看板
        .group_by(EquipmentModel.id, EquipmentModel.name, EquipmentModel.category, Device.ownership)
        .order_by(EquipmentModel.name)
    ).all()

    return [{
        "model_id": str(mid), "model_name": name, "category": category,
        "ownership": own,
        "total": total or 0, "rented": rented or 0,
        "available": available or 0, "pending": pending or 0,
    } for mid, name, category, own, total, rented, available, pending in rows]


def get_device_or_404(db: Session, device_id) -> Device:
    d = db.get(Device, device_id)
    if not d or d.deleted_at is not None:
        raise BusinessError("NOT_FOUND", "设备不存在", 404)
    return d


def update_device(db: Session, device_id, operator_id=None, **fields) -> Device:
    """白名单更新（schema DeviceUpdate 已约束字段集）；status/batch_id 不在其列。"""
    d = get_device_or_404(db, device_id)
    before = {k: str(getattr(d, k, None)) for k in fields}
    for k, v in fields.items():
        if v is not None and hasattr(d, k):
            setattr(d, k, v)
    db.flush()
    from app.services import audit_service as _audit
    _audit.log(db, user_id=operator_id, action="UPDATE", target_type="device",
               target_id=d.id, before_json=before,
               after_json={k: str(getattr(d, k, None)) for k in fields})
    # S3（缺陷#6）：预付款金额/日期变更 → 同步台账行（金额编辑同步 K7）
    from app.services import prepayment_service as _pp
    _pp.ensure_device_ledger(db, d, actor_id=operator_id)
    return d


def delete_device(db: Session, device_id, operator_id=None) -> Device:
    d = get_device_or_404(db, device_id)
    d.deleted_at = datetime.now(timezone.utc)
    db.flush()
    from app.services import audit_service as _audit
    _audit.log(db, user_id=operator_id, action="DELETE", target_type="device",
               target_id=d.id, before_json={"sn": d.sn})
    return d


def import_devices(db: Session, *, project_id, equipment_model_id, filebytes: bytes,
                   operator_id=None) -> int:
    """Excel 批量导入设备（复用 excel_service 模式）：每行建一台设备，缺省自动 SN。

    支持列（缺陷#11：中文表头优先，英文旧表头兼容）：
    SN 序列号/金租模式/单台月计费额(元)/单台采购原值(元)/预付款分摊(元)/权属
    """
    proj = db.get(Project, project_id)
    if not proj or proj.deleted_at is not None:
        raise BusinessError("NOT_FOUND", "项目不存在", 404)
    if not db.get(EquipmentModel, equipment_model_id):
        raise BusinessError("NOT_FOUND", "设备型号不存在", 404)
    # 缺陷#12：xls 老格式/损坏文件转友好错误（openpyxl 只支持 xlsx，BadZipFile 不再裸 500）
    try:
        wb = load_workbook(BytesIO(filebytes), read_only=True)
        ws = wb.active
        if ws is None:
            raise ValueError("工作簿无有效工作表")
        first_row = next(ws.iter_rows(min_row=1, max_row=1), None)
        if first_row is None:
            raise ValueError("空文件：无表头行")
    except BusinessError:
        raise
    except Exception as exc:  # noqa: BLE001  BadZipFile/KeyError/InvalidFileException 等
        raise BusinessError(
            "BAD_REQUEST",
            "文件无法解析：请上传 .xlsx 格式（Excel 2007 及以上），"
            "不支持 .xls 老格式；也不要把 .xls/.csv 改后缀成 .xlsx。"
            f"（解析失败原因：{type(exc).__name__}）",
            400,
        ) from exc
    headers = [str(c.value).strip() for c in first_row]
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(v is None for v in row):
            continue
        data = {}
        for i, h in enumerate(headers):
            field = _HEADER_MAP.get(h)
            if field and i < len(row) and row[i] not in (None, ""):
                data[field] = row[i]
        for k in ("monthly_price", "purchase_value", "prepayment_amount"):
            if k in data:
                data[k] = Decimal(str(data[k]))
        create_device(db, project_id=project_id, equipment_model_id=equipment_model_id,
                      operator_id=operator_id, **data)
        count += 1
    if count == 0:
        raise BusinessError(
            "BAD_REQUEST",
            "没有导入任何设备：请确认数据从第 2 行开始填写，且表头与模版一致"
            "（SN 序列号/金租模式/单台月计费额(元)/单台采购原值(元)/预付款分摊(元)/权属）",
            400,
        )
    return count
