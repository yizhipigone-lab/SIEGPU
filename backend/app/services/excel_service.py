"""Excel 导入导出（openpyxl）。导出多实体；导入仅支持全文本字段实体（suppliers/customers）。"""
from io import BytesIO

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.exceptions import BusinessError
from app.models.capital import CapitalTransaction
from app.models.master import Customer, Supplier

# key -> (model, columns, importable)
EXPORTS = {
    "suppliers": (Supplier, ["name", "type", "contact_person", "contact_phone", "bank_account", "notes"], True),
    "customers": (Customer, ["name", "industry", "contact_person", "contact_phone", "credit_rating", "notes"], True),
    "capital_transactions": (
        CapitalTransaction,
        ["transaction_date", "source_type", "direction", "amount", "category", "note"],
        False,
    ),
}


def export_xlsx(db: Session, key: str) -> BytesIO:
    if key not in EXPORTS:
        raise BusinessError("BAD_REQUEST", f"不支持导出：{key}", 400)
    model, cols, _ = EXPORTS[key]
    rows = db.execute(select(model)).scalars().all()
    wb = Workbook()
    ws = wb.active
    ws.title = key
    ws.append(cols)
    for r in rows:
        ws.append([getattr(r, c, None) for c in cols])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def import_xlsx(db: Session, key: str, filebytes: bytes) -> int:
    if key not in EXPORTS or not EXPORTS[key][2]:
        raise BusinessError("BAD_REQUEST", f"不支持导入：{key}（仅 suppliers/customers）", 400)
    model, cols, _ = EXPORTS[key]
    wb = load_workbook(BytesIO(filebytes), read_only=True)
    ws = wb.active
    headers = [str(c.value).strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(v is None for v in row):
            continue
        data = {}
        for i, h in enumerate(headers):
            if h in cols and i < len(row) and row[i] not in (None, ""):
                data[h] = row[i]
        if not data:
            continue
        db.add(model(**data))
        count += 1
    db.flush()
    return count
