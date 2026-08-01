"""报表 + Excel 测试。"""
import uuid
from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.models.master import Supplier
from app.services import excel_service as esvc
from app.services import report_service as rsvc
from app.services import capital_service as caps
from app.models.project import Project


def _user(db):
    from app.models.user import User
    u = User(username=f"u{uuid.uuid4().hex[:6]}", display_name="t", password_hash="x",
             role="FINANCE_DIRECTOR", active=True)
    db.add(u); db.flush(); return u


def _proj(db):
    p = Project(name="P", code=f"c{uuid.uuid4().hex[:6]}"); db.add(p); db.flush(); return p


def test_capital_monthly_aggregation(db):
    u = _user(db); p = _proj(db)
    caps.record_transaction(db, created_by=u.id, project_id=p.id, source_type="自有资金",
                            direction="IN", amount=Decimal("1000"), transaction_date=date(2026, 1, 5))
    caps.record_transaction(db, created_by=u.id, project_id=p.id, source_type="自有资金",
                            direction="IN", amount=Decimal("500"), transaction_date=date(2026, 1, 20))
    caps.record_transaction(db, created_by=u.id, project_id=p.id, source_type="自有资金",
                            direction="OUT", amount=Decimal("300"), transaction_date=date(2026, 2, 1))
    months = rsvc.capital_monthly(db)
    m = {x["month"]: x for x in months}
    assert m["2026-01"]["in"] == Decimal("1500") and m["2026-01"]["net"] == Decimal("1500")
    assert m["2026-02"]["out"] == Decimal("300") and m["2026-02"]["net"] == Decimal("-300")


def test_excel_export_import_roundtrip(db):
    db.add(Supplier(name="供应商1", type="设备供应商", contact_person="张三"))
    db.add(Supplier(name="供应商2", type="资金供应商", contact_person="李四"))
    db.flush()
    buf = esvc.export_xlsx(db, "suppliers")
    wb = load_workbook(buf)
    ws = wb.active
    assert ws.max_row == 3  # 1 header + 2 数据
    # 反向导入到空库（先软删？测试隔离由 fixture 回滚保证）
    out = BytesIO()
    wb2 = Workbook(); ws2 = wb2.active
    ws2.append(["name", "type", "contact_person"])
    ws2.append(["新供应商", "设备供应商", "王五"])
    wb2.save(out); out.seek(0)
    n = esvc.import_xlsx(db, "suppliers", out.getvalue())
    assert n == 1


def test_project_overview(db):
    u = _user(db); p = _proj(db)
    caps.record_transaction(db, created_by=u.id, project_id=p.id, source_type="自有资金",
                            direction="IN", amount=Decimal("800000"), transaction_date=date(2026, 1, 1))
    rows = rsvc.project_overview(db)
    row = [r for r in rows if r["project_id"] == str(p.id)][0]
    assert row["net_position"] == Decimal("800000")
    assert row["leasing_count"] == 0 and row["asset_count"] == 0
